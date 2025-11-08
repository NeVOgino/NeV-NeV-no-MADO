#!/usr/bin/env python3
"""
Extract data from the Excel file and create a structured JSON data file
"""
import openpyxl
import json

def extract_data():
    # Load the Excel file
    wb = openpyxl.load_workbook('NEVの窓_1110.xlsx')
    
    # Structure for storing data
    data = {
        '全員向け': {},
        '職員向け': {}
    }
    
    # Extract data from 全員向け sheet
    ws_all = wb['全員向け']
    data['全員向け'] = extract_all_staff_data(ws_all)
    
    # Extract data from 職員向け sheet
    ws_staff = wb['職員向け']
    data['職員向け'] = extract_staff_data(ws_staff)
    
    return data

def extract_all_staff_data(ws):
    """Extract data from 全員向け sheet"""
    result = {
        'information': [],
        '共通コーナー': [],
        '会議室予約': []
    }
    
    # Extract INFORMATION section (rows 12-40)
    for row in ws.iter_rows(min_row=12, max_row=40, values_only=True):
        if row[1]:  # If there's a date in column B
            entry = {
                'date': str(row[1]) if row[1] else '',
                'content': str(row[2]) if row[2] else '',
                'detail': str(row[3]) if row[3] else ''
            }
            if any(entry.values()):
                result['information'].append(entry)
    
    # Extract 共通コーナー section (column F)
    for row in ws.iter_rows(min_row=10, max_row=30, values_only=True):
        if row[5]:  # Column F
            item = str(row[5])
            if item and item not in result['共通コーナー']:
                result['共通コーナー'].append(item)
    
    # Extract 会議室予約 section (column H)
    for row in ws.iter_rows(min_row=9, max_row=20, values_only=True):
        if row[7]:  # Column H
            item = str(row[7])
            if item and item not in result['会議室予約']:
                result['会議室予約'].append(item)
    
    return result

def extract_staff_data(ws):
    """Extract data from 職員向け sheet with subdivisions"""
    result = {
        'information': [],
        '標準書コーナー': {
            '基本規則': [],
            '稟議・契約関係': [],
            '組織・運営': [],
            '就業規則・出張・旅費他': [],
            '安全衛生他': [],
            'ITセキュリティ関係他': [],
            'その他': []
        },
        '職員向け情報コーナー': [],
        '各種帳票フォーマット': []
    }
    
    # Extract INFORMATION section (rows 9-13)
    for row in ws.iter_rows(min_row=9, max_row=13, values_only=True):
        if row[1]:  # If there's a date in column B
            entry = {
                'date': str(row[1]) if row[1] else '',
                'content': str(row[2]) if row[2] else '',
                'detail': str(row[3]) if row[3] else ''
            }
            if entry['date'] and entry['date'] != '発信日':
                result['information'].append(entry)
    
    # Extract 標準書コーナー section (column F, rows 7-63)
    current_subdivision = '基本規則'
    for row in ws.iter_rows(min_row=7, max_row=63, values_only=True):
        if row[5]:  # Column F
            item = str(row[5]).strip()
            if not item:
                continue
            
            # Check if this is a subdivision header
            if item.startswith('【') and item.endswith('】'):
                # Extract subdivision name
                subdivision_name = item[1:-1]  # Remove 【 and 】
                if subdivision_name in result['標準書コーナー']:
                    current_subdivision = subdivision_name
            elif item != '標準書コーナー':
                # Add item to current subdivision
                if current_subdivision in result['標準書コーナー']:
                    result['標準書コーナー'][current_subdivision].append(item)
    
    # Extract 職員向け情報コーナー (rows 21-34)
    for row in ws.iter_rows(min_row=21, max_row=34, values_only=True):
        if row[2]:  # Column C
            item = str(row[2]).strip()
            if item and item != '職員向け情報コーナー':
                result['職員向け情報コーナー'].append(item)
    
    # Extract 各種帳票フォーマット (column H)
    for row in ws.iter_rows(min_row=9, max_row=15, values_only=True):
        if row[7]:  # Column H
            item = str(row[7]).strip()
            if item and not item.startswith('※'):
                result['各種帳票フォーマット'].append(item)
    
    return result

def main():
    print("Extracting data from Excel file...")
    data = extract_data()
    
    # Save to JSON file
    with open('data.json', 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    
    print("Data extracted successfully to data.json")
    
    # Print summary
    print("\n=== Data Summary ===")
    print(f"\n全員向け:")
    print(f"  - Information entries: {len(data['全員向け']['information'])}")
    print(f"  - 共通コーナー items: {len(data['全員向け']['共通コーナー'])}")
    print(f"  - 会議室予約 items: {len(data['全員向け']['会議室予約'])}")
    
    print(f"\n職員向け:")
    print(f"  - Information entries: {len(data['職員向け']['information'])}")
    print(f"  - 標準書コーナー subdivisions:")
    for subdivision, items in data['職員向け']['標準書コーナー'].items():
        print(f"    - {subdivision}: {len(items)} items")
    print(f"  - 職員向け情報コーナー items: {len(data['職員向け']['職員向け情報コーナー'])}")
    print(f"  - 各種帳票フォーマット items: {len(data['職員向け']['各種帳票フォーマット'])}")

if __name__ == '__main__':
    main()
